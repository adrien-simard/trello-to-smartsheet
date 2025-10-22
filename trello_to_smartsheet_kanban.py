#!/usr/bin/env python3
"""
Trello to Smartsheet Kanban Migration Script

This script imports a Trello board (from JSON export) into Smartsheet,
preserving the Kanban board structure with lists as lanes, cards as rows,
and comments as native Smartsheet discussions.

Requirements:
    - Python 3.9+
    - smartsheet-python-sdk
    - Smartsheet API token (set in environment variable SMARTSHEET_ACCESS_TOKEN)

Usage:
    python trello_to_smartsheet_kanban.py <trello_export.json>
"""

import json
import os
import sys
import time
from datetime import datetime
from typing import Dict, List, Optional, Any

import smartsheet
from smartsheet.models import Sheet, Column, Row, Cell, Discussion, Comment

# Import version info
try:
    from __version__ import __version__
except ImportError:
    __version__ = "1.0.0"

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class TrelloToSmartsheetMigrator:
    """Main class for migrating Trello boards to Smartsheet"""

    def __init__(self, api_token: str, folder_id: Optional[int] = None, email_mapping_file: Optional[str] = None):
        """
        Initialize the migrator with Smartsheet API credentials.

        Args:
            api_token: Smartsheet API access token
            folder_id: Optional folder ID to create sheet in
            email_mapping_file: Optional Excel file with member name to email mapping
        """
        self.smartsheet_client = smartsheet.Smartsheet(api_token)
        self.smartsheet_client.errors_as_exceptions(True)
        self.folder_id = folder_id

        # Load email mapping if provided, otherwise use empty dict (auto-generate emails)
        if email_mapping_file:
            self.email_mapping = self.load_email_mapping(email_mapping_file)
        else:
            self.email_mapping = {}
            print("[*] No email mapping file provided - emails will be auto-generated from names")

    def load_trello_data(self, file_path: str) -> Dict[str, Any]:
        """
        Load and parse Trello JSON export file.

        Args:
            file_path: Path to Trello JSON export file

        Returns:
            Parsed Trello board data

        Raises:
            FileNotFoundError: If file doesn't exist
            json.JSONDecodeError: If file is not valid JSON
        """
        print(f"[*] Loading Trello data from: {file_path}")

        with open(file_path, 'r', encoding='utf-8') as f:
            data = json.load(f)

        print(f"[OK] Loaded board: {data.get('name', 'Unknown')}")
        print(f"   Lists: {len(data.get('lists', []))}")
        print(f"   Cards: {len(data.get('cards', []))}")
        print(f"   Actions: {len(data.get('actions', []))}")

        return data

    def load_email_mapping(self, file_path: str) -> Dict[str, str]:
        """
        Load email mapping from Excel file.

        Args:
            file_path: Path to Excel file with name -> email mapping

        Returns:
            Dictionary mapping member names to email addresses
        """
        if not OPENPYXL_AVAILABLE:
            print("[WARN] openpyxl not available, cannot load email mapping")
            return {}

        if not os.path.exists(file_path):
            print(f"[WARN] Email mapping file not found: {file_path}")
            return {}

        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active

            email_map = {}
            # Skip header row
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] and row[1]:  # Both name and email present
                    name = str(row[0]).strip()
                    email = str(row[1]).strip()
                    email_map[name] = email
                    # Also store lowercase version for flexible matching
                    email_map[name.lower()] = email

            print(f"[OK] Loaded {len(email_map) // 2} email mappings from {file_path}")
            return email_map

        except Exception as e:
            print(f"[WARN] Failed to load email mapping: {e}")
            return {}

    def create_smartsheet_columns(self) -> List[Column]:
        """
        Create the column structure for Smartsheet.

        Returns:
            List of Column objects matching the spec
        """
        columns = [
            Column({
                'title': 'Card Name',
                'type': 'TEXT_NUMBER',
                'primary': True
            }),
            Column({
                'title': 'List',
                'type': 'PICKLIST',
                'options': []  # Will be populated dynamically
            }),
            Column({
                'title': 'Description',
                'type': 'TEXT_NUMBER'
            }),
            Column({
                'title': 'Due Date',
                'type': 'DATE'
            }),
            Column({
                'title': 'Members',
                'type': 'MULTI_CONTACT_LIST'
            }),
            Column({
                'title': 'Labels',
                'type': 'MULTI_PICKLIST',
                'options': []  # Will be populated dynamically with all label names
            }),
            Column({
                'title': 'URL',
                'type': 'TEXT_NUMBER'
            }),
            Column({
                'title': 'Created Date',
                'type': 'DATE'
            })
        ]

        return columns

    def extract_list_names(self, trello_data: Dict[str, Any]) -> List[str]:
        """
        Extract unique list names from Trello data for dropdown options.

        Args:
            trello_data: Parsed Trello board data

        Returns:
            List of unique list names in order
        """
        lists = trello_data.get('lists', [])
        # Filter out archived lists and maintain order
        active_lists = [lst for lst in lists if not lst.get('closed', False)]
        return [lst['name'] for lst in active_lists]

    def extract_label_names(self, trello_data: Dict[str, Any]) -> List[str]:
        """
        Extract unique label names from Trello data for multi-select dropdown options.

        Args:
            trello_data: Parsed Trello board data

        Returns:
            List of unique label names
        """
        labels = trello_data.get('labels', [])
        label_names = set()

        for label in labels:
            # Use label name if available, otherwise use color
            label_name = label.get('name', '').strip()
            if not label_name:
                label_name = label.get('color', 'Unlabeled')
            if label_name:
                label_names.add(label_name)

        # Return sorted list for consistency
        return sorted(list(label_names))

    def create_sheet(self, board_name: str, list_names: List[str], label_names: List[str]) -> Sheet:
        """
        Create a new Smartsheet sheet with proper column structure.

        Args:
            board_name: Name of the Trello board
            list_names: List of Trello list names for dropdown
            label_names: List of Trello label names for multi-select dropdown

        Returns:
            Created Sheet object
        """
        sheet_name = f"Trello Import - {board_name}"
        # Smartsheet limits sheet names to 50 characters
        if len(sheet_name) > 50:
            # Truncate board name to fit within limit
            max_board_name_length = 50 - len("Trello Import - ")
            truncated_board_name = board_name[:max_board_name_length]
            sheet_name = f"Trello Import - {truncated_board_name}"
        print(f"\n[*] Creating Smartsheet: {sheet_name}")

        # Create columns
        columns = self.create_smartsheet_columns()

        # Set List and Labels column options
        for col in columns:
            if col.title == 'List':
                col.options = list_names
            elif col.title == 'Labels':
                col.options = label_names

        # Create sheet specification
        sheet_spec = Sheet({
            'name': sheet_name,
            'columns': columns
        })

        # Create the sheet (in folder if specified)
        if self.folder_id:
            response = self.smartsheet_client.Folders.create_sheet_in_folder(
                self.folder_id,
                sheet_spec
            )
        else:
            response = self.smartsheet_client.Home.create_sheet(sheet_spec)

        # The response already contains the sheet with all details including column IDs
        sheet = response.result

        print(f"[OK] Sheet created with ID: {sheet.id}")

        return sheet

    def build_card_lookup(self, trello_data: Dict[str, Any]) -> Dict[str, Dict[str, Any]]:
        """
        Build a lookup dictionary for Trello cards by ID.

        Args:
            trello_data: Parsed Trello board data

        Returns:
            Dictionary mapping card IDs to card objects
        """
        return {card['id']: card for card in trello_data.get('cards', [])}

    def build_list_lookup(self, trello_data: Dict[str, Any]) -> Dict[str, str]:
        """
        Build a lookup dictionary for Trello list names by ID.

        Args:
            trello_data: Parsed Trello board data

        Returns:
            Dictionary mapping list IDs to list names
        """
        return {lst['id']: lst['name'] for lst in trello_data.get('lists', [])}

    def build_member_lookup(self, trello_data: Dict[str, Any]) -> Dict[str, Dict[str, str]]:
        """
        Build a lookup dictionary for Trello member details by ID.

        Args:
            trello_data: Parsed Trello board data

        Returns:
            Dictionary mapping member IDs to member info (name, email)
        """
        members = {}
        for member in trello_data.get('members', []):
            member_id = member['id']
            full_name = member.get('fullName', member.get('username', 'Unknown'))

            # Try to get email from mapping file first
            email = self.email_mapping.get(full_name)
            if not email:
                # Try case-insensitive match
                email = self.email_mapping.get(full_name.lower())
            if not email:
                # Fallback: generate email from full name (prenom.nom@epfl.ch)
                email = self.generate_email_from_name(full_name)

            members[member_id] = {
                'name': full_name,
                'email': email
            }
        return members

    def generate_email_from_name(self, full_name: str) -> str:
        """
        Generate an EPFL email address from a full name.

        Args:
            full_name: Person's full name (e.g., "Adrien SIMARD")

        Returns:
            Email address in format prenom.nom@epfl.ch
        """
        # Clean and split the name
        name_parts = full_name.strip().split()

        if len(name_parts) == 0:
            return "unknown@epfl.ch"
        elif len(name_parts) == 1:
            # Only one name part
            return f"{name_parts[0].lower()}@epfl.ch"
        else:
            # First name is typically first, last name is last
            first_name = name_parts[0].lower()
            last_name = name_parts[-1].lower()
            return f"{first_name}.{last_name}@epfl.ch"

    def build_label_lookup(self, trello_data: Dict[str, Any]) -> Dict[str, str]:
        """
        Build a lookup dictionary for Trello label names by ID.

        Args:
            trello_data: Parsed Trello board data

        Returns:
            Dictionary mapping label IDs to label names
        """
        labels = {}
        for label in trello_data.get('labels', []):
            label_id = label['id']
            label_name = label.get('name', label.get('color', 'Unlabeled'))
            labels[label_id] = label_name
        return labels

    def parse_trello_date(self, date_str: Optional[str]) -> Optional[str]:
        """
        Parse Trello ISO date string to Smartsheet date format (YYYY-MM-DD).

        Args:
            date_str: ISO 8601 date string from Trello

        Returns:
            Date string in YYYY-MM-DD format, or None
        """
        if not date_str:
            return None

        try:
            dt = datetime.fromisoformat(date_str.replace('Z', '+00:00'))
            return dt.strftime('%Y-%m-%d')
        except (ValueError, AttributeError):
            return None

    def create_row_from_card(
        self,
        card: Dict[str, Any],
        column_map: Dict[str, int],
        list_lookup: Dict[str, str],
        member_lookup: Dict[str, Dict[str, str]],
        label_lookup: Dict[str, str]
    ) -> Row:
        """
        Create a Smartsheet row from a Trello card.

        Args:
            card: Trello card object
            column_map: Dictionary mapping column names to column IDs
            list_lookup: Dictionary mapping list IDs to names
            member_lookup: Dictionary mapping member IDs to names
            label_lookup: Dictionary mapping label IDs to names

        Returns:
            Row object ready to be added to Smartsheet
        """
        # Build cells
        cells = []

        # Card Name (primary column)
        cells.append(Cell({
            'column_id': column_map['Card Name'],
            'value': card.get('name', 'Untitled')
        }))

        # List (dropdown)
        list_id = card.get('idList')
        list_name = list_lookup.get(list_id, '')
        cells.append(Cell({
            'column_id': column_map['List'],
            'value': list_name
        }))

        # Description
        description = card.get('desc', '')
        if description:
            cells.append(Cell({
                'column_id': column_map['Description'],
                'value': description
            }))

        # Due Date
        due_date = self.parse_trello_date(card.get('due'))
        if due_date:
            cells.append(Cell({
                'column_id': column_map['Due Date'],
                'value': due_date
            }))

        # Members (multi-contact list)
        member_ids = card.get('idMembers', [])
        if member_ids:
            contacts = []
            for mid in member_ids:
                member_info = member_lookup.get(mid)
                if member_info:
                    contacts.append({
                        'objectType': 'CONTACT',
                        'name': member_info['name'],
                        'email': member_info['email']
                    })

            if contacts:
                cells.append(Cell({
                    'column_id': column_map['Members'],
                    'object_value': {
                        'objectType': 'MULTI_CONTACT',
                        'values': contacts
                    }
                }))

        # Labels (multi-select)
        label_ids = card.get('idLabels', [])
        if label_ids:
            label_names = [label_lookup.get(lid, 'Unknown') for lid in label_ids]
            # For MULTI_PICKLIST, set objectValue with proper format
            cells.append(Cell({
                'column_id': column_map['Labels'],
                'object_value': {
                    'objectType': 'MULTI_PICKLIST',
                    'values': label_names
                }
            }))

        # URL
        card_url = card.get('shortUrl', card.get('url', ''))
        if card_url:
            cells.append(Cell({
                'column_id': column_map['URL'],
                'value': card_url
            }))

        # Created Date
        created_date = self.parse_trello_date(card.get('dateLastActivity'))
        if created_date:
            cells.append(Cell({
                'column_id': column_map['Created Date'],
                'value': created_date
            }))

        return Row({
            'cells': cells
        })

    def add_cards_to_sheet(
        self,
        sheet: Sheet,
        trello_data: Dict[str, Any]
    ) -> Dict[str, int]:
        """
        Add all Trello cards as rows to the Smartsheet.

        Args:
            sheet: Smartsheet Sheet object
            trello_data: Parsed Trello board data

        Returns:
            Dictionary mapping Trello card IDs to Smartsheet row IDs
        """
        print(f"\n[*] Adding cards to sheet...")

        # Build lookups
        list_lookup = self.build_list_lookup(trello_data)
        member_lookup = self.build_member_lookup(trello_data)
        label_lookup = self.build_label_lookup(trello_data)

        # Build column map (name -> ID)
        column_map = {col.title: col.id for col in sheet.columns}

        # Create rows from cards
        cards = [card for card in trello_data.get('cards', []) if not card.get('closed', False)]
        rows = []

        for card in cards:
            row = self.create_row_from_card(
                card,
                column_map,
                list_lookup,
                member_lookup,
                label_lookup
            )
            rows.append(row)

        # Add rows to sheet in batches
        if rows:
            response = self.smartsheet_client.Sheets.add_rows(sheet.id, rows)
            added_rows = response.result

            print(f"[OK] Added {len(added_rows)} cards")

            # Build card ID -> row ID mapping
            card_to_row_map = {}
            for i, row in enumerate(added_rows):
                card_to_row_map[cards[i]['id']] = row.id

            return card_to_row_map

        return {}

    def extract_comments_for_cards(
        self,
        trello_data: Dict[str, Any]
    ) -> Dict[str, List[Dict[str, Any]]]:
        """
        Extract comment actions from Trello data, grouped by card ID.

        Args:
            trello_data: Parsed Trello board data

        Returns:
            Dictionary mapping card IDs to lists of comment objects
        """
        comments_by_card = {}

        for action in trello_data.get('actions', []):
            if action.get('type') == 'commentCard':
                card_id = action.get('data', {}).get('card', {}).get('id')
                if card_id:
                    if card_id not in comments_by_card:
                        comments_by_card[card_id] = []

                    # Get member creator info
                    member_creator = action.get('memberCreator', {})
                    author_name = member_creator.get('fullName', 'Unknown')
                    member_id = member_creator.get('id')

                    comments_by_card[card_id].append({
                        'text': action.get('data', {}).get('text', ''),
                        'author_name': author_name,
                        'member_id': member_id,
                        'date': action.get('date', '')
                    })

        return comments_by_card

    def add_comments_to_rows(
        self,
        sheet_id: int,
        trello_data: Dict[str, Any],
        card_to_row_map: Dict[str, int]
    ):
        """
        Add Trello comments as Smartsheet discussions on rows.

        Args:
            sheet_id: Smartsheet sheet ID
            trello_data: Parsed Trello board data
            card_to_row_map: Dictionary mapping card IDs to row IDs
        """
        print(f"\n[*] Adding comments as discussions...")

        # Build member lookup for comment authors
        member_lookup = self.build_member_lookup(trello_data)
        comments_by_card = self.extract_comments_for_cards(trello_data)

        total_comments = 0

        for card_id, comments in comments_by_card.items():
            row_id = card_to_row_map.get(card_id)
            if not row_id:
                continue

            for comment_data in comments:
                # Get comment details
                text = comment_data['text']
                author_name = comment_data['author_name']
                member_id = comment_data.get('member_id')
                date_str = comment_data['date']

                # Get author email from member lookup
                author_email = None
                if member_id and member_id in member_lookup:
                    author_email = member_lookup[member_id]['email']

                # Parse date for display
                date_display = ''
                if date_str:
                    try:
                        dt = datetime.fromisoformat(date_str.replace('Z', '+00:00'))
                        date_display = dt.strftime('%Y-%m-%d %H:%M')
                    except:
                        date_display = date_str

                # Format: [Author (email) - Date]\nComment text
                formatted_text = f"[{author_name}"
                if author_email:
                    formatted_text += f" ({author_email})"
                if date_display:
                    formatted_text += f" - {date_display}"
                formatted_text += f"]\n{text}"

                # Create discussion with comment
                discussion = Discussion()
                discussion.comment = Comment()
                discussion.comment.text = formatted_text

                try:
                    self.smartsheet_client.Discussions.create_discussion_on_row(
                        sheet_id,
                        row_id,
                        discussion
                    )
                    total_comments += 1
                except Exception as e:
                    print(f"[WARN] Failed to add comment to row {row_id}: {e}")

        print(f"[OK] Added {total_comments} comments")

    def migrate_board(self, trello_file_path: str) -> int:
        """
        Main migration workflow: load Trello data and create Smartsheet.

        Args:
            trello_file_path: Path to Trello JSON export file

        Returns:
            Smartsheet sheet ID
        """
        # Load Trello data
        trello_data = self.load_trello_data(trello_file_path)

        # Extract board name, list names, and label names
        board_name = trello_data.get('name', 'Untitled Board')
        list_names = self.extract_list_names(trello_data)
        label_names = self.extract_label_names(trello_data)

        # Create Smartsheet
        sheet = self.create_sheet(board_name, list_names, label_names)

        # Add cards as rows
        card_to_row_map = self.add_cards_to_sheet(sheet, trello_data)

        # Add comments as discussions
        self.add_comments_to_rows(sheet.id, trello_data, card_to_row_map)

        print(f"\n[SUCCESS] Migration complete!")
        print(f"   Sheet ID: {sheet.id}")
        print(f"   Sheet name: {sheet.name}")
        print(f"\n[NEXT STEPS]")
        print(f"   1. Open the sheet in Smartsheet")
        print(f"   2. Switch to Card View")
        print(f"   3. Set 'List' as the lane field")
        print(f"   4. Enjoy your Kanban board!")

        return sheet.id


def main():
    """Main entry point for the script."""

    print(f"Trello to Smartsheet Migration Tool v{__version__}")
    print("=" * 60)

    # Check arguments
    if len(sys.argv) < 2:
        print("Usage: python trello_to_smartsheet_kanban.py <trello_export.json> [api_token] [folder_id] [email_mapping.xlsx]")
        sys.exit(1)

    trello_file = sys.argv[1]

    # Check file exists
    if not os.path.exists(trello_file):
        print(f"Error: File not found: {trello_file}")
        sys.exit(1)

    # Get API token from command line or environment
    api_token = None
    if len(sys.argv) >= 3:
        api_token = sys.argv[2]
    else:
        api_token = os.getenv('SMARTSHEET_ACCESS_TOKEN')

    if not api_token:
        print("Error: SMARTSHEET_ACCESS_TOKEN not provided")
        print("\nTo set it:")
        print("  Environment: set SMARTSHEET_ACCESS_TOKEN=your_token_here")
        print("  Or pass as argument: python trello_to_smartsheet_kanban.py file.json your_token_here")
        sys.exit(1)

    # Get optional folder ID
    folder_id = None
    if len(sys.argv) >= 4:
        try:
            folder_id = int(sys.argv[3])
            print(f"[*] Will create sheet in folder ID: {folder_id}")
        except ValueError:
            print(f"[WARN] Invalid folder ID: {sys.argv[3]}, creating in Home")

    # Get optional email mapping file
    email_mapping_file = None
    if len(sys.argv) >= 5:
        email_mapping_file = sys.argv[4]
        if not os.path.exists(email_mapping_file):
            print(f"[WARN] Email mapping file not found: {email_mapping_file}")
            email_mapping_file = None

    # Run migration
    try:
        migrator = TrelloToSmartsheetMigrator(api_token, folder_id, email_mapping_file)
        migrator.migrate_board(trello_file)
    except Exception as e:
        print(f"\n[ERROR] Migration failed: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == '__main__':
    main()
